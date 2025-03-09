import openpyxl
import logging
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from .forms import FileUploadForm
from .models import Startup, StartupApplication
from datetime import datetime
import re
from django.core.paginator import Paginator
from decimal import Decimal, InvalidOperation

logger = logging.getLogger(__name__)

def upload_file(request):
    """Handle file upload and initiate processing based on file type."""
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.save()
            if not uploaded_file.file.name.endswith('.xlsx'):
                messages.error(request, "Only .xlsx files are supported.")
                return redirect('file_upload_success')
            try:
                wb = openpyxl.load_workbook(uploaded_file.file.path, data_only=True)
                sheet = wb.active

                required_columns = ["Application ID"]
                header_row_idx, detected_headers = detect_header_row(sheet, required_columns)

                if header_row_idx is None:
                    messages.error(request, "Could not detect a valid header row in the file.")
                    return redirect('file_upload_success')

                if "application id" in [h.lower() for h in detected_headers]:
                    success = process_pva_file(uploaded_file.file.path)
                else:
                    success = process_excel_file(uploaded_file.file.path)

                if success:
                    uploaded_file.processed = True
                    uploaded_file.save()
                    messages.success(request, "File uploaded and processed successfully!")
                else:
                    messages.error(request, "File processing failed. Please check the format.")

            except Exception as e:
                messages.error(request, f"Error reading file: {e}")

            return redirect('file_upload_success')

    else:
        form = FileUploadForm()

    return render(request, 'upload.html', {'form': form})

def process_pva_file(file_path):
    """Process a PVA Excel file and store application data in the database."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # Find header row by searching for 'Application ID'
        header_row_idx = None
        headers = []
        for i in range(1, 10):
            row_values = [str(cell.value).strip() if cell.value else "" for cell in sheet[i]]
            if "Application ID" in row_values:
                header_row_idx = i
                headers = row_values
                break

        if not header_row_idx:
            return False

        column_indices = {header.lower(): idx for idx, header in enumerate(headers) if header}

        required_columns = ['application id', 'startup/person name']
        if any(col not in column_indices for col in required_columns):
            return False

        data_start_row = header_row_idx + 1
        success_count = 0

        def safe_get_value(idx, row_data):
            """Safely retrieve a cell value, returning None if out of bounds or empty."""
            return str(row_data[idx]).strip() if idx is not None and idx < len(row_data) and row_data[idx] else None

        for row_idx in range(data_start_row, sheet.max_row + 1):
            row_data = [cell.value for cell in sheet[row_idx]]
            if not any(row_data):
                continue

            try:
                application_id = safe_get_value(column_indices.get('application id'), row_data)
                startup_name = safe_get_value(column_indices.get('startup/person name'), row_data)
                applicant_name = safe_get_value(column_indices.get('primary contact name'), row_data)
                primary_contact_title = safe_get_value(column_indices.get('primary contact title'), row_data)
                email = safe_get_value(column_indices.get('primary contact email address'), row_data)
                phone = safe_get_value(column_indices.get('phone'), row_data)
                brief_description = safe_get_value(column_indices.get('brief description'), row_data)
                business_stage = safe_get_value(column_indices.get('fund stage'), row_data)
                average_score = safe_get_value(column_indices.get('average score'), row_data)

                contact_info = {
                    "city": safe_get_value(column_indices.get('city'), row_data),
                    "country": safe_get_value(column_indices.get('country'), row_data),
                    "website": safe_get_value(column_indices.get('website'), row_data),
                    "facebook": safe_get_value(column_indices.get('facebook'), row_data),
                    "twitter": safe_get_value(column_indices.get('twitter'), row_data),
                    "linkedin": safe_get_value(column_indices.get('linkedin'), row_data),
                    "github": safe_get_value(column_indices.get('github'), row_data),
                    "additional_links": [
                        safe_get_value(column_indices.get('additional link 1'), row_data),
                        safe_get_value(column_indices.get('additional link 2'), row_data),
                        safe_get_value(column_indices.get('additional link 3'), row_data)
                    ]
                }

                videos = {
                    "product_video": safe_get_value(column_indices.get('product video'), row_data),
                    "team_video": safe_get_value(column_indices.get('team video'), row_data)
                }

                registration_info = {
                    "status": safe_get_value(column_indices.get('are you registered or incorporated?'), row_data),
                    "location": safe_get_value(column_indices.get('where are you registered or incorporated?'), row_data),
                    "start_date": safe_get_value(column_indices.get('when did you start this company?'), row_data)
                }

                progress_status = {
                    "description": safe_get_value(column_indices.get('what do you do in detail?'), row_data),
                    "differentiators": safe_get_value(column_indices.get("what's different/interesting about your startup?"), row_data),
                    "stage": safe_get_value(column_indices.get('how far along are you?'), row_data)
                }

                financials = {
                    "funds_raised": safe_get_value(column_indices.get('how much money raised since start?'), row_data),
                    "runway": safe_get_value(column_indices.get('how much runway do you have left?'), row_data),
                    "raising": safe_get_value(column_indices.get('raising'), row_data),
                    "amount_raising": safe_get_value(column_indices.get('amount raising'), row_data),
                    "currency": safe_get_value(column_indices.get('amount currency'), row_data),
                    "valuation": safe_get_value(column_indices.get('valuation'), row_data),
                    "valuation_currency": safe_get_value(column_indices.get('valuation currency'), row_data)
                }

                customer_info = {
                    "markets": safe_get_value(column_indices.get('skills or markets'), row_data).split(',') if safe_get_value(column_indices.get('skills or markets'), row_data) else [],
                    "customers": safe_get_value(column_indices.get('key customers/users?'), row_data).split(',') if safe_get_value(column_indices.get('key customers/users?'), row_data) else []
                }

                if not application_id or not startup_name:
                    continue

                startup = Startup.objects.filter(item_name__icontains=startup_name).first()
                if not startup:
                    continue

                application, created = StartupApplication.objects.update_or_create(
                    application_id=application_id,
                    defaults={
                        "startup": startup,
                        "applicant_name": applicant_name,
                        "primary_contact_title": primary_contact_title,
                        "email": email,
                        "phone": phone,
                        "brief_description": brief_description,
                        "business_stage": business_stage,
                        "average_score": float(average_score) if average_score else None,
                        "contact_info": contact_info,
                        "videos": videos,
                        "registration_info": registration_info,
                        "progress_status": progress_status,
                        "financials": financials,
                        "customer_info": customer_info,
                    }
                )
                success_count += 1

            except Exception as e:
                logger.error(f"Error processing row {row_idx}: {str(e)}")

        return success_count > 0

    except Exception as e:
        logger.error(f"Critical error processing PVA file: {str(e)}")
        return False

def parse_date(value):
    """Convert Excel dates or string dates into Python date objects."""
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, str):
        value = value.strip()
        match = re.match(r"([a-zA-Z]+)\s*'(\d{2})", value)
        if match:
            month_name, year_short = match.groups()
            try:
                month = datetime.strptime(month_name, "%b").month
                year = int("20" + year_short)
                return datetime(year, month, 1).date()
            except ValueError:
                return None

        formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y"]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue

    return None

def parse_funding(value):
    """Extract currency and numeric value from a funding string (e.g., 'GBP 550000')."""
    if not value or not isinstance(value, str):
        return None, None

    match = re.match(r"([A-Za-z]+)\s*([\d,]+(?:\.\d{1,2})?)", value.strip())
    if match:
        currency, amount = match.groups()
        amount = amount.replace(',', '')
        return currency.upper(), float(amount)

    return None, None

def detect_header_row(sheet, required_columns, pva_columns=None, threshold=0.4):
    """Detect the header row by matching required column names."""
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_values = [str(cell).strip() if cell else "" for cell in row]
        non_empty_cells = sum(1 for cell in row_values if cell)
        if non_empty_cells < 4:
            continue

        matching_pipeline_cols = [col for col in required_columns if col.lower() in map(str.lower, row_values)]
        matching_pva_cols = [col for col in (pva_columns or []) if col.lower() in map(str.lower, row_values)]

        if len(matching_pipeline_cols) >= len(required_columns) * threshold:
            return i, row_values
        if len(matching_pva_cols) >= len((pva_columns or [])) * threshold:
            return i, row_values

    return None, None

def process_excel_file(file_path):
    """Process a Pipeline Excel file and save startup data to the database."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        required_columns = ["Startup ID", "Item Name", "Pipeline", "Location", "Markets",
                            "Founder 1 Name", "Founder 1 Role", "Founder 2 Name", "Founder 2 Role",
                            "Founder 3 Name", "Founder 3 Role", "Founder 4 Name", "Founder 4 Role",
                            "Founder 5 Name", "Founder 5 Role", "Founder 6 Name", "Founder 6 Role",
                            "Founder 7 Name", "Founder 7 Role",
                            "Website 1", "AngelList", "Linkedin", "GitHub", "Twitter", "Facebook", "Google Plus",
                            "Tagline", "Milestone", "Revenue Model", "Source 1 Name", "Source 1 Type",
                            "Last Contact", "Files", "Incorporated", "Founded", "Differentiators",
                            "Description", "Interfaces", "Total Funding", "Cash Runway", "Clients",
                            "Videos", "Rev Last 12 Months", "Rev Last Month", "Rounds"]

        header_row_index, headers = detect_header_row(sheet, required_columns)
        if not header_row_index:
            return False

        column_mapping = {col.lower(): i for i, col in enumerate(headers) if col}

        for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            if not any(row):
                continue

            startup_id = str(row[column_mapping.get('startup id', '')])
            item_name = str(row[column_mapping.get('item name', '')])
            pipeline = row[column_mapping.get('pipeline', '')]
            location = row[column_mapping.get('location', '')]
            markets = row[column_mapping.get('markets', '')]

            founders = []
            for i in range(1, 8):
                name = row[column_mapping.get(f'founder {i} name'.lower(), '')]
                role = row[column_mapping.get(f'founder {i} role'.lower(), '')]
                if name and role:
                    founders.append({'name': name.strip(), 'role': role.strip()})

            social_media = {
                'website': row[column_mapping.get('website 1', '')],
                'angellist': row[column_mapping.get('angellist', '')],
                'linkedin': row[column_mapping.get('linkedin', '')],
                'github': row[column_mapping.get('github', '')],
                'twitter': row[column_mapping.get('twitter', '')],
                'facebook': row[column_mapping.get('facebook', '')],
                'google_plus': row[column_mapping.get('google plus', '')]
            }
            social_media = {k: v.strip() for k, v in social_media.items() if v and v.strip()}

            tagline = row[column_mapping.get('tagline', '')]
            milestone = row[column_mapping.get('milestone', '')]
            revenue_model = row[column_mapping.get('revenue model', '')]
            last_contact = parse_date(row[column_mapping.get('last contact', '')])
            incorporated = row[column_mapping.get('incorporated', '')]
            founded_date = None
            if 'founded' in column_mapping:
                raw_founded = row[column_mapping.get('founded', '')]
                founded_date = parse_date(raw_founded) if raw_founded else None

            differentiators = row[column_mapping.get('differentiators', '')]
            description = row[column_mapping.get('description', '')]
            interfaces = row[column_mapping.get('interfaces', '')]
            funding_value = row[column_mapping.get('total funding', '')] or ""
            total_funding_currency, total_funding_amount = parse_funding(funding_value)
            cash_runway = str(row[column_mapping.get('cash runway', '')] or "").strip()
            rev_last_12_months = str(row[column_mapping.get('rev last 12 months', '')] or "").strip()
            rev_last_month = str(row[column_mapping.get('rev last month', '')] or "").strip()
            rounds = int(row[column_mapping.get('rounds', '')] or 0)
            clients = row[column_mapping.get('clients', '')]

            startup_data = {
                'startup_id': startup_id, 'item_name': item_name, 'pipeline': pipeline,
                'location': location, 'markets': markets, 'founders': founders,
                'social_media': social_media, 'tagline': tagline, 'milestone': milestone,
                'revenue_model': revenue_model, 'last_contact': last_contact,
                'incorporated': incorporated, 'founded_date': founded_date,
                'differentiators': differentiators, 'description': description,
                'interfaces': interfaces, 'total_funding_currency': total_funding_currency,
                'total_funding_amount': total_funding_amount, 'cash_runway': cash_runway,
                'rev_last_12_months': rev_last_12_months, 'rev_last_month': rev_last_month,
                'rounds': rounds, 'clients': clients
            }

            startup, created = Startup.objects.update_or_create(
                startup_id=startup_id, defaults=startup_data
            )

        return True

    except Exception as e:
        return False

def file_upload_success(request):
    """Render the success page after file upload."""
    return render(request, 'upload_success.html')

def homepage(request):
    """Render the homepage."""
    return render(request, 'homepage.html')

def startup_list(request):
    """Display a paginated list of startups with filtering and sorting."""
    search_query = request.GET.get('q', '').strip()
    min_funding = request.GET.get('min_funding')
    max_funding = request.GET.get('max_funding')

    startups = Startup.objects.all().prefetch_related('applications')
    if search_query:
        startups = startups.filter(item_name__icontains=search_query)

    try:
        if min_funding and min_funding != 'None':
            min_funding = Decimal(min_funding)
            startups = startups.filter(total_funding_amount__gte=min_funding)
    except (ValueError, InvalidOperation):
        pass

    try:
        if max_funding and max_funding != 'None':
            max_funding = Decimal(max_funding)
            startups = startups.filter(total_funding_amount__lte=max_funding)
    except (ValueError, InvalidOperation):
        pass

    allowed_sort_fields = ['item_name', 'location', 'total_funding_amount', 'created_at']
    sort_by = request.GET.get('sort', 'item_name')
    order = request.GET.get('order', 'asc')

    if sort_by not in allowed_sort_fields:
        sort_by = 'item_name'

    if order == 'desc':
        sort_by = f"-{sort_by}"

    startups = startups.order_by(sort_by)

    startups_with_dash = []
    for startup in startups:
        application = startup.applications.first()
        score = application.average_score if (application and application.average_score is not None) else 0
        dash_length = (score / 5.0) * 282.7
        startups_with_dash.append({
            'startup': startup,
            'dash_length': dash_length
        })

    paginator = Paginator(startups_with_dash, 10)
    page_number = request.GET.get('page')
    startups_page = paginator.get_page(page_number)

    return render(request, 'startups/startup_list.html', {
        'startups': startups_page,
        'sort_by': sort_by.lstrip('-'),
        'order': order,
        'search_query': search_query,
        'min_funding': min_funding if min_funding != 'None' else '',
        'max_funding': max_funding if max_funding != 'None' else ''
    })

def startup_detail(request, startup_id):
    """Display details for a specific startup."""
    startup = get_object_or_404(Startup, id=startup_id)
    return render(request, 'startups/startup_detail.html', {'startup': startup})

def calculate_quality_score(data):
    """Calculate a completeness score (0-100) based on key fields."""
    required_fields = [data.get('applicant_name'), data.get('email'), data.get('phone'),
                       data.get('brief_description'), data.get('business_stage')]
    total = len(required_fields)
    completed = sum(1 for field in required_fields if field and field.strip())
    return (completed / total) * 100 if total > 0 else 0